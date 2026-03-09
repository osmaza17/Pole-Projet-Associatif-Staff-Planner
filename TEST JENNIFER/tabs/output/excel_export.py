"""
Excel exporter for the Output tab.

Generates a multi-sheet .xlsx workbook with two views per day:
  • Task View  — tasks on rows, hours on columns, cells = people names
  • People View — people on rows, hours on columns, cells = task labels
"""

import flet as ft
from constants import TASK_COLORS


class ExcelExporter:
    """
    Self-contained helper that manages the FilePicker and openpyxl export.

    Usage
    -----
    >>> exporter = ExcelExporter(page)
    >>> # … later, on button click:
    >>> exporter.export(sol, people, tasks, hours, days,
    ...                 availability, emergency, person_colors)
    """

    def __init__(self, page: ft.Page):
        self._page = page
        self._pending: dict = {}
        self._picker = ft.FilePicker(on_result=self._on_save_result)
        page.overlay.append(self._picker)

    # ── Public API ─────────────────────────────────────────────────────

    def export(self, sol, people, tasks, hours, days,
               availability, emergency, person_colors):
        """Store context and open the save-file dialog."""
        self._pending = dict(
            sol=sol, people=people, tasks=tasks, hours=hours,
            days=days, availability=availability,
            emergency=emergency, person_colors=person_colors)
        self._picker.save_file(
            dialog_title="Export grid to Excel",
            file_name="staffing_output.xlsx",
            allowed_extensions=["xlsx"])

    # ── FilePicker callback ────────────────────────────────────────────

    def _on_save_result(self, e: ft.FilePickerResultEvent):
        """Write the openpyxl workbook once the user picks a path."""
        if not e.path:
            return

        d             = self._pending
        sol           = d.get("sol", {})
        people        = d.get("people", [])
        tasks         = d.get("tasks", [])
        hours         = d.get("hours", {})
        days          = d.get("days", [])
        availability  = d.get("availability", {})
        emergency     = d.get("emergency", {})
        person_colors = d.get("person_colors", {})
        assignment    = sol.get("assignment", {})

        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        # ── Shared style helpers ───────────────────────────────────────
        def _fill(hex6: str) -> PatternFill:
            return PatternFill("solid", fgColor=hex6.lstrip("#").upper())

        def _border(color="B0BEC5") -> Border:
            s = Side(style="thin", color=color.lstrip("#").upper())
            return Border(left=s, right=s, top=s, bottom=s)

        def _font(bold=False, color="000000", size=10) -> Font:
            return Font(bold=bold, color=color.lstrip("#").upper(), size=size)

        def _align(h="center", v="center", indent=0) -> Alignment:
            return Alignment(horizontal=h, vertical=v,
                             indent=indent, wrap_text=False)

        def _style(cell, fill=None, font=None, align=None, border=None):
            if fill   is not None: cell.fill      = fill
            if font   is not None: cell.font      = font
            if align  is not None: cell.alignment  = align
            if border is not None: cell.border    = border

        HDR_FILL   = _fill("263238")
        EMPTY_FILL = _fill("D3D3D3")
        TOTAL_FILL = _fill("546E7A")
        EVEN_FILL  = _fill("ECEFF1")
        BORDER     = _border()
        HDR_BORDER = _border("455A64")

        tc = {t: TASK_COLORS[i % len(TASK_COLORS)]
              for i, t in enumerate(tasks)}

        # ════════════════════════════════════════════════════════════════
        # PASS 1 — TASK VIEW sheets  (one per day)
        # ════════════════════════════════════════════════════════════════
        for day in days:
            if day not in assignment:
                continue
            asgn    = assignment[day]
            day_hrs = hours[day]
            base    = str(day)[:28]

            ws = wb.create_sheet(title=f"{base} - Tasks")
            ws.freeze_panes = "B2"
            total_col = len(day_hrs) + 2

            # Header row
            ws.cell(1, 1, "Task")
            for h_idx, h in enumerate(day_hrs, start=2):
                ws.cell(1, h_idx, str(h))
            ws.cell(1, total_col, "Total")
            for col in range(1, total_col + 1):
                _style(ws.cell(1, col),
                       fill   = HDR_FILL,
                       font   = _font(bold=True, color="FFFFFF", size=11),
                       align  = _align("left" if col == 1 else "center",
                                       indent=1 if col == 1 else 0),
                       border = HDR_BORDER)
            ws.row_dimensions[1].height = 24

            cur_row = 2
            for t in tasks:
                t_bg_hex, t_fg_hex = tc[t]
                TASK_FILL      = _fill(t_bg_hex)
                task_font_name = _font(bold=True, color=t_fg_hex, size=11)

                pph = {
                    h: [p for p in people if asgn.get(p, {}).get(h) == t]
                    for h in day_hrs
                }
                total_ppl = sum(len(v) for v in pph.values())
                n_rows    = max(1, max((len(v) for v in pph.values()),
                                       default=0))

                # Hour cells
                for h_idx, h in enumerate(day_hrs, start=2):
                    for row_off, person in enumerate(pph[h]):
                        cell  = ws.cell(cur_row + row_off, h_idx, person)
                        p_hex = person_colors.get(person, t_fg_hex)
                        _style(cell,
                               fill   = TASK_FILL,
                               font   = _font(bold=True, color=p_hex, size=10),
                               align  = _align("left", indent=1),
                               border = BORDER)
                    for row_off in range(len(pph[h]), n_rows):
                        _style(ws.cell(cur_row + row_off, h_idx),
                               fill=EMPTY_FILL, border=BORDER)

                # Task name cell — merged vertically
                if n_rows > 1:
                    ws.merge_cells(start_row=cur_row, start_column=1,
                                   end_row=cur_row + n_rows - 1,
                                   end_column=1)
                _style(ws.cell(cur_row, 1, t),
                       fill   = TASK_FILL,
                       font   = task_font_name,
                       align  = _align("left", "center", indent=1),
                       border = BORDER)

                # Total cell — merged vertically
                if n_rows > 1:
                    ws.merge_cells(start_row=cur_row, start_column=total_col,
                                   end_row=cur_row + n_rows - 1,
                                   end_column=total_col)
                _style(ws.cell(cur_row, total_col,
                               total_ppl if total_ppl else ""),
                       fill   = TOTAL_FILL,
                       font   = _font(bold=True, color="FFFFFF", size=11),
                       align  = _align("center", "center"),
                       border = BORDER)
                for row_off in range(1, n_rows):
                    ws.cell(cur_row + row_off, total_col).border = BORDER

                for row_off in range(n_rows):
                    ws.row_dimensions[cur_row + row_off].height = 18
                cur_row += n_rows

            # Column widths
            ws.column_dimensions["A"].width = 14
            for h_idx in range(2, len(day_hrs) + 2):
                ws.column_dimensions[get_column_letter(h_idx)].width = 14
            ws.column_dimensions[get_column_letter(total_col)].width = 8

        # ════════════════════════════════════════════════════════════════
        # PASS 2 — PEOPLE VIEW sheets  (one per day)
        # ════════════════════════════════════════════════════════════════
        for day in days:
            if day not in assignment:
                continue
            asgn    = assignment[day]
            day_hrs = hours[day]
            base    = str(day)[:28]

            ws2 = wb.create_sheet(title=f"{base} - People")
            ws2.freeze_panes = "B2"
            total_col2 = len(day_hrs) + 2

            # Header row
            ws2.cell(1, 1, "Person")
            for h_idx, h in enumerate(day_hrs, start=2):
                ws2.cell(1, h_idx, str(h))
            ws2.cell(1, total_col2, "Total")
            for col in range(1, total_col2 + 1):
                _style(ws2.cell(1, col),
                       fill   = HDR_FILL,
                       font   = _font(bold=True, color="FFFFFF", size=11),
                       align  = _align("left" if col == 1 else "center",
                                       indent=1 if col == 1 else 0),
                       border = HDR_BORDER)
            ws2.row_dimensions[1].height = 24

            for idx_p, p in enumerate(people):
                r     = idx_p + 2
                fill  = EVEN_FILL if idx_p % 2 == 0 else _fill("FFFFFF")
                total = 0

                ws2.cell(r, 1, p)
                for h_idx, h in enumerate(day_hrs, start=2):
                    task_val = asgn.get(p, {}).get(h, "")
                    ws2.cell(r, h_idx, task_val)
                    if task_val:
                        total += 1
                        t_bg, t_fg = tc.get(task_val, ("#FFFFFF", "#000000"))
                        _style(ws2.cell(r, h_idx),
                               fill   = _fill(t_bg),
                               font   = _font(bold=True, color=t_fg, size=10),
                               align  = _align("center"),
                               border = BORDER)
                    else:
                        _style(ws2.cell(r, h_idx),
                               fill=fill, align=_align("center"),
                               border=BORDER)

                ws2.cell(r, total_col2, total)
                p_hex = person_colors.get(p, "000000")
                _style(ws2.cell(r, 1),
                       fill   = fill,
                       font   = _font(bold=True, color=p_hex, size=10),
                       align  = _align("left", indent=1),
                       border = BORDER)
                _style(ws2.cell(r, total_col2),
                       fill   = TOTAL_FILL,
                       font   = _font(bold=True, color="FFFFFF", size=10),
                       align  = _align("center"),
                       border = BORDER)
                ws2.row_dimensions[r].height = 18

            # Column widths
            ws2.column_dimensions["A"].width = 18
            for h_idx in range(2, len(day_hrs) + 2):
                ws2.column_dimensions[get_column_letter(h_idx)].width = 14
            ws2.column_dimensions[get_column_letter(total_col2)].width = 8

        # ── Save ───────────────────────────────────────────────────────
        path = e.path if e.path.endswith(".xlsx") else e.path + ".xlsx"
        try:
            wb.save(path)
            self._page.snack_bar = ft.SnackBar(
                content=ft.Text(f"✔  Exported → {path}"),
                bgcolor=ft.Colors.GREEN_700, duration=4000)
            self._page.snack_bar.open = True
        except Exception as exc:
            self._page.snack_bar = ft.SnackBar(
                content=ft.Text(f"✘  Export failed: {exc}"),
                bgcolor=ft.Colors.RED_700, duration=6000)
            self._page.snack_bar.open = True
        self._page.update()